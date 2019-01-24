#!/usr/bin/env python3
# -*-coding:utf-8-*-

import os

#chp8
# 倒斜杠\，正斜杠/
print(os.path.join('usr','bin','weidong')) # windows 下返回\ Linux中返回/
print(os.getcwd()) # 获取当前目录
os.chdir('d:\\') # 设置当前目录，Windows中要用\\
print(os.getcwd())
# os.makedir 创建新文件夹
print(os.path.abspath(os.getcwd())) # 判断是否是绝对路径：os.path.isabs()
print(os.path.relpath(os.getcwd(),'D:\\10000h\\0408')) # 返回一个相对路径
print(os.path.dirname('D:\\10000h\\0408')) # 返回最后一个反斜杠前的字符串，
print(os.path.basename('D:\\10000h\\0408')) # 返回最后一个斜杠后的所有内容
print(os.path.split('D:\\10000h\\0408')) # 返回目录名称和基本名称，return元组
path = 'D:\\10000h\\0408'
path.split(os.path.sep) # 路径分隔符os.path.sep

print(os.listdir('D:\\10000h')) # 返回目录文件列表
print(os.path.getsize('D:\\workspace\\python\\backup\\20180928\\learning_1.py')) #查看文件大小

# 检查路径有效性
print(os.path.exists('D:\\10000h')) # 是否存在
print(os.path.isfile('jjj')) # 是否是文件
print(os.path.isdir('jkl')) # 是否是目录

# 读取文件
helloFile = open('D:\\workspace\\python\\python_quick_start\\chp.txt') # 打开文件，传入w开启写模式
content = helloFile.read()
contentNew = helloFile.readline()
print(content)
print(contentNew)

# chp9
import shutil
import os
import send2trash
import zipfile

os.chdir('D:\\workspace\\python\\backup\\20181214')
shutil.copy('D:\\workspace\\python\\backup\\20181214\\study_sorted.py','D:\\workspace\\python\\python_quick_start')
# 第二个参数可以指定文件名

# shutil.move(source,destination) 目标文件夹必须存在
shutil.move('D:\\workspace\\python\\backup\\20181214\\move.txt','D:\\workspace\\python\\python_quick_start\\newMove.txt')

os.unlink('D:\\workspace\\testdir\\test.txt') # 删除此处文件
os.rmdir('D:\\workspace\\testdir') # 删除此处目录，文件夹下必须没有文件
shutil.rmtree('D:\\workspace\\testdir\\lib2') #删除文件夹
send2trash.send2trash('D:\\workspace\\testdir\\test2.txt') #删除文件，送入回收站

for folderName,subfolders,filenames in os.walk('D:\\'):
    print('The current folder is ' + folderName)
    for subfolder in subfolders:
        print('SUBFOLDER OF ' + folderName + ': ' + subfolder)
    for filename in filenames:
        print('FILE INSIDE ' + folderName + ': '+ filename)
    print('') # os.walk 返回三个值，1.当前文件夹目录名，2.当前文件夹的子文件夹列表 3.当前文件夹下文件列表


# chp10
import traceback
import os
import logging

logging.basicConfig(level=logging.DEBUG,format='%(asctime)s - %(levelname)s - %(message)s')

def minus(num):
    if num ==0:
        raise ZeroDivisionError("num can't be 0")
    else:
        print(20/num)

# 通常是调用该函数的代码知道如何处理异常，而不是该函数本身。所以你常常会看到raise语句在一
# 个函数中，try和except语句在调用该函数的代码中
os.chdir('D:\\workspace\\python\\python_quick_start')
try:
 minus(0)
except ZeroDivisionError as e:
    print('入参不能为0')

# 收集异常日志
try:
    minus(0)
except:
    errorFile = open('errorFile.txt','w')
    logging.debug('创建文件')
    logging.disable(logging.CRITICAL)
    errorFile.write(traceback.format_exc()) # 禁用日志
    logging.debug('写入traceback信息')
    errorFile.close()
    print('已收集异常日志')

# logging 模块 ：你可以随心所欲地在程序中想加多少就加多少，稍后只要加入一次logging.disable
# logging level

# chp11
import webbrowser
import requests
# webbrowser.open('https://www.baidu.com')
res = requests.get('http://www.zhixue.com')
print(res.status_code)
#newRes = requests.get('http://www.zhixue.com/unkonwpage.html')
#newRes.raise_for_status() # 当访问出错抛出异常

aRes = requests.get('http://www.gutenberg.org/cache/epub/1112/pg1112.txt')
try:
    aRes.raise_for_status()
    playFile = open('RomeoAndJuliet.txt', 'wb')
    for chunk in aRes.iter_content(100000):
        playFile.write(chunk) # 指定每次读取内容大小
    playFile.close()
except Exception as e:
    print('error: %s' % (e))

# chp12
import openpyxl
import os
os.chdir('D:\\workspace\\python\\python_quick_start')
wb = openpyxl.load_workbook('example.xlsx')
print(type(wb)) # 获取workbook对象
print(wb.get_sheet_names())
print(wb.get_sheet_by_name('Sheet3'))
print(type(wb.get_sheet_by_name('Sheet3')))
sheet = wb.get_active_sheet() # 获取活动的sheet页
print(sheet) 
print(sheet['A1']) # 获取单元格对象Cell
print(sheet['A1'].value)
c = sheet['B2']
print(c.value)
print('c的坐标：第%s行，%s列' %(c.row,c.column)) # 打印坐标
d = sheet.cell(row=1,column=2)
print(d)
print(d.value)
for i in range(1,3): # range不会取末尾值，
    print(i,sheet.cell(row=i,column =2).value)

# 列字母和数字转化
# from openpyxl.cell import get_column_letter,column_index_from_string #???

# 获取活动页行和列
# print(sheet.get_highest_row()) # 报错了？是因为活动页当前没有边界？
# print(sheet.get_highest_column())

# 获得行和列
a = tuple(sheet['A1':'B2'])
for rowOfCellObjects in a:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate,cellObj.value)